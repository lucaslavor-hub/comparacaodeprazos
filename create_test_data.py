#!/usr/bin/env python3
import openpyxl
from datetime import datetime, timedelta
import random

# Create Seven file (relatório)
wb_seven = openpyxl.Workbook()
ws_seven = wb_seven.active
ws_seven.title = "Publicações"

# Headers for Seven
seven_headers = [
    "Número Processo",
    "Conteúdo",
    "Data Diário",
    "Nome Encontrado",
    "Cliente",
    "UF",
    "Status Publicação"
]

for col, header in enumerate(seven_headers, 1):
    ws_seven.cell(1, col, header)

# Add sample data for Seven
seven_data = [
    ["1234567-89.2024.1.02.1234", "Conteúdo do processo 1", datetime(2024, 1, 15), "João Silva", "Cliente A", "SP", "Publicado"],
    ["1234568-89.2024.1.02.1235", "Conteúdo do processo 2", datetime(2024, 1, 16), "Maria Santos", "Cliente B", "RJ", "Publicado"],
    ["1234569-89.2024.1.02.1236", "Conteúdo do processo 3", datetime(2024, 1, 17), "João Silva", "Cliente C", "MG", "Publicado"],
    ["1234570-89.2024.1.02.1237", "Conteúdo do processo 4", datetime(2024, 1, 18), "Pedro Costa", "Cliente D", "BA", "Publicado"],
    ["1234571-89.2024.1.02.1238", "Conteúdo do processo 5", datetime(2024, 1, 19), "Ana Oliveira", "Cliente E", "SP", "Publicado"],
]

for row_idx, row_data in enumerate(seven_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_seven.cell(row_idx, col_idx, value)

wb_seven.save(r"c:\Users\lucas\Desktop\comparacaodeprazos\public\seven_sample.xlsx")
print("✅ Arquivo Seven criado: seven_sample.xlsx")

# Create Serur file (planilha de referência)
wb_serur = openpyxl.Workbook()
ws_serur = wb_serur.active
ws_serur.title = "Processos"

# Headers for Serur
serur_headers = [
    "Número Processo",
    "Conteúdo",
    "Data Publicação",
    "Cliente",
    "UF",
]

for col, header in enumerate(serur_headers, 1):
    ws_serur.cell(1, col, header)

# Add sample data for Serur (some matching, some not)
serur_data = [
    ["1234567-89.2024.1.02.1234", "Conteúdo do processo 1", datetime(2024, 1, 15), "Cliente A", "SP"],
    ["1234568-89.2024.1.02.1235", "Conteúdo do processo 2", datetime(2024, 1, 16), "Cliente B", "RJ"],
    ["1234572-89.2024.1.02.1240", "Conteúdo do processo extra", datetime(2024, 1, 20), "Cliente F", "BA"],
]

for row_idx, row_data in enumerate(serur_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_serur.cell(row_idx, col_idx, value)

wb_serur.save(r"c:\Users\lucas\Desktop\comparacaodeprazos\public\serur_sample.xlsx")
print("✅ Arquivo Serur criado: serur_sample.xlsx")

print("\n📊 Dados de teste criados com sucesso!")
